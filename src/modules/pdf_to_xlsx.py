import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment


class PdfToXlsxConverter:
    """Convert PDF tables to XLSX format."""

    def __init__(self, input_path: str):
        self.input_path = input_path

    def convert(self, output_path: str, extract_all_pages: bool = True) -> dict:
        """Convert PDF tables to XLSX.
        
        Args:
            output_path: Path to save XLSX file
            extract_all_pages: Extract tables from all pages
        
        Returns:
            dict with conversion stats
        """
        import os
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        table_count = 0
        
        with pdfplumber.open(self.input_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.extract_tables()
                
                if not tables:
                    continue
                
                for table_idx, table in enumerate(tables, 1):
                    table_count += 1
                    sheet_name = f"Page{page_num}_Table{table_idx}"
                    ws = wb.create_sheet(title=sheet_name)
                    
                    for row_idx, row in enumerate(table, 1):
                        for col_idx, cell in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=cell)
                            if row_idx == 1:
                                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)
                                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')
                
                if not extract_all_pages:
                    break
        
        if table_count == 0:
            ws = wb.create_sheet(title="No Tables Found")
            ws.cell(row=1, column=1, value="No tables detected in PDF")
        
        wb.save(output_path)
        
        return {
            "input_file": self.input_path,
            "output_file": output_path,
            "output_size": os.path.getsize(output_path),
            "table_count": table_count
        }

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass
